import openpyxl

work_book = openpyxl.load_workbook("demowebshopdata.xlsx")
sheet = work_book["valid_creds"]
rows = sheet.max_row
columns = sheet.max_column

valid_creds = []
for row in range(2, rows + 1):
    nested_list = []
    for column in range(1, columns + 1):
        data = sheet.cell(row, column).value
        nested_list.append(data)
    valid_creds.append(nested_list)

print(valid_creds)