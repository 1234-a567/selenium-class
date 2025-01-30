import openpyxl

def get_data_from_excel(filename, sheetname):
    work_book = openpyxl.load_workbook(filename)
    sheet = work_book[sheetname]
    rows = sheet.max_row
    columns = sheet.max_column

    valid_creds = []
    for row in range(2, rows + 1):
        nested_list = []
        for column in range(1, columns + 1):
            data = sheet.cell(row, column).value
            nested_list.append(data)
        valid_creds.append(nested_list)
    return valid_creds

def get_invalid_data_from_excel(filename, sheetname):
    work_book = openpyxl.load_workbook(filename)
    sheet = work_book[sheetname]
    rows = sheet.max_row
    columns = sheet.max_column

    valid_creds = []
    for row in range(2, rows + 1):
        nested_list = []
        for column in range(1, columns + 1):
            data = sheet.cell(row, column).value
            nested_list.append(data)
        valid_creds.append(nested_list)
    return valid_creds

def get_passwords_from_excel(filename, sheetname):
    work_book = openpyxl.load_workbook(filename)
    sheet = work_book[sheetname]
    rows = sheet.max_row
    passwords = []
    for row in range(2, rows + 1):
        password = sheet.cell(row, 1).value
        passwords.append(password)

    return passwords

def get_emails_from_excel(filename, sheetname):
    work_book = openpyxl.load_workbook(filename)
    sheet = work_book[sheetname]
    rows = sheet.max_row
    passwords = []
    for row in range(2, rows + 1):
        password = sheet.cell(row, 1).value
        passwords.append(password)

    return passwords
